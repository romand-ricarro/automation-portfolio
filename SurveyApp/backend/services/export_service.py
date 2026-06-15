"""
Export service for generating PDF and Excel reports.
Provides functions to export session analysis data in various formats.
"""
import io
import re
import logging
from datetime import datetime
from typing import Dict, Any, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, ListFlowable, ListItem
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models.database import Session, QuestionAnalysis, CommonIssue, ActionItem

logger = logging.getLogger('insightpulse.export_service')


def strip_markdown(text: str) -> str:
    """
    Convert markdown text to plain text by removing markdown formatting.

    Args:
        text: Markdown formatted text

    Returns:
        Plain text with markdown formatting removed
    """
    if not text:
        return text

    # Remove headers (### Header -> Header)
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)

    # Remove bold (**text** or __text__ -> text)
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'__(.+?)__', r'\1', text)

    # Remove italic (*text* or _text_ -> text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'_(.+?)_', r'\1', text)

    # Remove inline code (`code` -> code)
    text = re.sub(r'`(.+?)`', r'\1', text)

    # Remove links [text](url) -> text
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)

    # Remove bullet points (- item or * item -> item)
    text = re.sub(r'^[\-\*]\s+', '• ', text, flags=re.MULTILINE)

    # Remove numbered lists (1. item -> item)
    text = re.sub(r'^\d+\.\s+', '', text, flags=re.MULTILINE)

    # Clean up multiple newlines
    text = re.sub(r'\n{3,}', '\n\n', text)

    return text.strip()


def get_session_export_data(session: Session) -> Dict[str, Any]:
    """
    Gather all session data needed for export.

    Args:
        session: The Session object to export

    Returns:
        Dictionary containing all export data
    """
    # Get analyses
    analyses = QuestionAnalysis.query.filter_by(session_id=session.id).all()

    # Get common issues
    common_issues = CommonIssue.query.filter_by(session_id=session.id).order_by(CommonIssue.display_order).all()

    # Get action items
    action_items = ActionItem.query.filter_by(session_id=session.id).all()

    # Get ratings
    ratings = session.ratings.to_dict() if session.ratings else None

    return {
        'session': session.to_dict(),
        'analyses': [a.to_dict() for a in analyses],
        'common_issues': [ci.to_dict() for ci in common_issues],
        'action_items': [ai.to_dict() for ai in action_items],
        'ratings': ratings,
    }


def generate_pdf_report(session: Session) -> bytes:
    """
    Generate a PDF report for a session.

    Args:
        session: The Session object to export

    Returns:
        PDF file as bytes
    """
    data = get_session_export_data(session)

    # Create buffer
    buffer = io.BytesIO()

    # Create document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )

    # Styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='CenterTitle',
        parent=styles['Heading1'],
        alignment=TA_CENTER,
        spaceAfter=30
    ))
    styles.add(ParagraphStyle(
        name='SectionTitle',
        parent=styles['Heading2'],
        spaceBefore=20,
        spaceAfter=10
    ))
    styles.add(ParagraphStyle(
        name='SubSection',
        parent=styles['Heading3'],
        spaceBefore=15,
        spaceAfter=8
    ))

    # Build content
    content = []

    # === PAGE 1: Cover ===
    content.append(Spacer(1, 2*inch))
    content.append(Paragraph("InsightPulse", styles['CenterTitle']))
    content.append(Paragraph("Session Analysis Report", styles['CenterTitle']))
    content.append(Spacer(1, inch))

    # Session info
    session_info = data['session']
    session_name = session_info.get('session_name') or session_info['session_id']
    content.append(Paragraph(f"<b>Session Name:</b> {session_name}", styles['Normal']))
    content.append(Spacer(1, 12))
    content.append(Paragraph(f"<b>Session ID:</b> {session_info['session_id']}", styles['Normal']))
    content.append(Spacer(1, 12))
    content.append(Paragraph(f"<b>Date:</b> {session_info['session_date']}", styles['Normal']))
    content.append(Spacer(1, 12))
    content.append(Paragraph(f"<b>Facilitator:</b> {session_info['facilitator_name'] or 'N/A'}", styles['Normal']))
    content.append(Spacer(1, 12))
    content.append(Paragraph(f"<b>Responses:</b> {session_info['num_responses'] or 0}", styles['Normal']))
    content.append(Spacer(1, 12))
    content.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))

    content.append(PageBreak())

    # === PAGE 2: Ratings Summary ===
    if data['ratings']:
        content.append(Paragraph("Session Ratings Summary", styles['SectionTitle']))

        ratings = data['ratings']
        rating_data = [
            ['Metric', 'Score'],
            ['Facilitator Understanding', f"{ratings.get('facilitator_understanding', 'N/A'):.2f}" if ratings.get('facilitator_understanding') else 'N/A'],
            ['Learning Mechanics', f"{ratings.get('learning_mechanics', 'N/A'):.2f}" if ratings.get('learning_mechanics') else 'N/A'],
            ['Q&A Support', f"{ratings.get('qa_support', 'N/A'):.2f}" if ratings.get('qa_support') else 'N/A'],
            ['Problem Articulation', f"{ratings.get('problem_articulation', 'N/A'):.2f}" if ratings.get('problem_articulation') else 'N/A'],
            ['Session Pace', f"{ratings.get('session_pace', 'N/A'):.2f}" if ratings.get('session_pace') else 'N/A'],
            ['Tools Helpfulness', f"{ratings.get('tools_helpfulness', 'N/A'):.2f}" if ratings.get('tools_helpfulness') else 'N/A'],
            ['Repeatability', f"{ratings.get('repeatability', 'N/A'):.2f}" if ratings.get('repeatability') else 'N/A'],
            ['Learning Objectives', f"{ratings.get('learning_objectives', 'N/A'):.2f}" if ratings.get('learning_objectives') else 'N/A'],
            ['Overall Quality', f"{ratings.get('overall_quality', 'N/A'):.2f}" if ratings.get('overall_quality') else 'N/A'],
        ]

        rating_table = Table(rating_data, colWidths=[4*inch, 1.5*inch])
        rating_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        content.append(rating_table)
        content.append(PageBreak())

    # === PAGES 3-N: Question Analyses ===
    if data['analyses']:
        content.append(Paragraph("Question Analyses", styles['SectionTitle']))

        for analysis in data['analyses']:
            content.append(Paragraph(analysis['question_text'], styles['SubSection']))

            # Strip markdown and wrap the analysis text
            analysis_text = strip_markdown(analysis['analysis_text'])
            analysis_text = analysis_text.replace('\n', '<br/>')
            content.append(Paragraph(analysis_text, styles['Normal']))
            content.append(Spacer(1, 20))

        content.append(PageBreak())

    # === Common Issues Table ===
    if data['common_issues']:
        content.append(Paragraph("Common Issues & Themes", styles['SectionTitle']))

        issues_data = [['#', 'Issue', 'Evidence/Signal']]
        for idx, issue in enumerate(data['common_issues'], 1):
            issues_data.append([
                str(idx),
                Paragraph(issue['common_issue'], styles['Normal']),
                Paragraph(issue['evidence_signal'], styles['Normal'])
            ])

        issues_table = Table(issues_data, colWidths=[0.5*inch, 2.5*inch, 3*inch])
        issues_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        content.append(issues_table)
        content.append(PageBreak())

    # === Action Items ===
    if data['action_items']:
        content.append(Paragraph("Action Items", styles['SectionTitle']))

        actions_data = [['Issue', 'Action', 'Priority', 'Status', 'Assignee', 'Deadline']]
        for item in data['action_items']:
            actions_data.append([
                Paragraph(item['issue'][:50] + '...' if len(item['issue']) > 50 else item['issue'], styles['Normal']),
                Paragraph(item['action'][:50] + '...' if len(item['action']) > 50 else item['action'], styles['Normal']),
                item['priority'] or 'N/A',
                item['status'] or 'Open',
                item['person_in_charge'] or 'Unassigned',
                item['deadline'] or 'No deadline'
            ])

        actions_table = Table(actions_data, colWidths=[1.2*inch, 1.5*inch, 0.7*inch, 0.8*inch, 1*inch, 0.8*inch])
        actions_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        content.append(actions_table)

    # Build PDF
    doc.build(content)

    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(session: Session) -> bytes:
    """
    Generate an Excel report for a session.

    Args:
        session: The Session object to export

    Returns:
        Excel file as bytes
    """
    data = get_session_export_data(session)

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def auto_width(ws, min_width=12, max_width=50):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    # === Sheet 1: Overview ===
    ws_overview = wb.active
    ws_overview.title = "Overview"

    session_info = data['session']
    overview_data = [
        ['Session Overview', ''],
        ['Session ID', session_info['session_id']],
        ['Session Name', session_info['session_name'] or 'N/A'],
        ['Date', session_info['session_date']],
        ['Facilitator', session_info['facilitator_name'] or 'N/A'],
        ['Responses', session_info['num_responses'] or 0],
        ['Status', session_info['status']],
        ['Analyzed At', session_info['analyzed_at'] or 'Not analyzed'],
        ['', ''],
        ['Ratings', 'Score'],
    ]

    if data['ratings']:
        ratings = data['ratings']
        overview_data.extend([
            ['Facilitator Understanding', ratings.get('facilitator_understanding', 'N/A')],
            ['Learning Mechanics', ratings.get('learning_mechanics', 'N/A')],
            ['Q&A Support', ratings.get('qa_support', 'N/A')],
            ['Problem Articulation', ratings.get('problem_articulation', 'N/A')],
            ['Session Pace', ratings.get('session_pace', 'N/A')],
            ['Tools Helpfulness', ratings.get('tools_helpfulness', 'N/A')],
            ['Repeatability', ratings.get('repeatability', 'N/A')],
            ['Learning Objectives', ratings.get('learning_objectives', 'N/A')],
            ['Overall Quality', ratings.get('overall_quality', 'N/A')],
        ])

    for row_idx, row_data in enumerate(overview_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1 or row_idx == 10:
                cell.font = Font(bold=True)

    auto_width(ws_overview)

    # === Sheet 2: Question Analyses ===
    ws_analyses = wb.create_sheet("Question Analyses")

    ws_analyses.cell(row=1, column=1, value="Question")
    ws_analyses.cell(row=1, column=2, value="Analysis")
    style_header_row(ws_analyses, 1, 2)

    for row_idx, analysis in enumerate(data['analyses'], 2):
        ws_analyses.cell(row=row_idx, column=1, value=analysis['question_text']).alignment = cell_alignment
        # Strip markdown from analysis text for cleaner Excel output
        clean_analysis = strip_markdown(analysis['analysis_text'])
        ws_analyses.cell(row=row_idx, column=2, value=clean_analysis).alignment = cell_alignment
        ws_analyses.cell(row=row_idx, column=1).border = thin_border
        ws_analyses.cell(row=row_idx, column=2).border = thin_border

    ws_analyses.column_dimensions['A'].width = 40
    ws_analyses.column_dimensions['B'].width = 80

    # === Sheet 3: Common Issues ===
    ws_issues = wb.create_sheet("Common Issues")

    headers = ['#', 'Common Issue', 'Evidence/Signal']
    for col_idx, header in enumerate(headers, 1):
        ws_issues.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_issues, 1, len(headers))

    for row_idx, issue in enumerate(data['common_issues'], 2):
        ws_issues.cell(row=row_idx, column=1, value=issue.get('display_order', row_idx - 1)).border = thin_border
        ws_issues.cell(row=row_idx, column=2, value=issue['common_issue']).border = thin_border
        ws_issues.cell(row=row_idx, column=3, value=issue['evidence_signal']).border = thin_border
        for col in range(1, 4):
            ws_issues.cell(row=row_idx, column=col).alignment = cell_alignment

    ws_issues.column_dimensions['A'].width = 5
    ws_issues.column_dimensions['B'].width = 40
    ws_issues.column_dimensions['C'].width = 50

    # === Sheet 4: Action Items ===
    ws_actions = wb.create_sheet("Action Items")

    headers = ['Issue', 'Action', 'Priority', 'Status', 'Assignee', 'Deadline', 'Notes', 'Created']
    for col_idx, header in enumerate(headers, 1):
        ws_actions.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_actions, 1, len(headers))

    for row_idx, item in enumerate(data['action_items'], 2):
        ws_actions.cell(row=row_idx, column=1, value=item['issue']).border = thin_border
        ws_actions.cell(row=row_idx, column=2, value=item['action']).border = thin_border
        ws_actions.cell(row=row_idx, column=3, value=item['priority'] or 'N/A').border = thin_border
        ws_actions.cell(row=row_idx, column=4, value=item['status'] or 'Open').border = thin_border
        ws_actions.cell(row=row_idx, column=5, value=item['person_in_charge'] or 'Unassigned').border = thin_border
        ws_actions.cell(row=row_idx, column=6, value=item['deadline'] or 'No deadline').border = thin_border
        ws_actions.cell(row=row_idx, column=7, value=item['notes'] or '').border = thin_border
        ws_actions.cell(row=row_idx, column=8, value=item['created_at']).border = thin_border
        for col in range(1, 9):
            ws_actions.cell(row=row_idx, column=col).alignment = cell_alignment

    auto_width(ws_actions)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
